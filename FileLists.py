#"""Module used to search through OS files"""
import os
#Module used to write to xlsx sheet
import openpyxl

# Create an excel workbook with 3 sheets
workbook = openpyxl.Workbook()
sheet_txt = workbook.create_sheet("txt")
sheet_json = workbook.create_sheet("json")
sheet_vdf = workbook.create_sheet("vdf")

# Delete the default empty "Sheet" sheet
sheet = workbook["Sheet"]
workbook.remove(sheet)

# Define a list of file extensions to search for
extensions = [".txt", ".json", ".vdf"]

# Define a dictionary to map file extensions to sheets
sheets = {
    ".txt": sheet_txt,
    ".json": sheet_json,
    ".vdf": sheet_vdf,
}

# Add a header row to each sheet
sheet_txt.append(["file"])
sheet_json.append(["file"])
sheet_vdf.append(["file"])

# Travel through directories and subdirectories to search for files with the defined extensions
# Change the root directory to what you need

#Variable used to sort out unnecessary files used in line 42
fileReduction = ['brazilian','bulgarian','czech','danish','dutch','english','finnish','french','german','greek','hungarian','italian','japanese','korean','koreana','latam','norwegian','polish','portuguese','romanian','russian','schinese','spanish','swedish','tchinese','thai','turkish','ukrainian','vietnamese', 'controller', 'readme', 'gamepad', 'chord', 'bigpicture', 'desktop','disabled', 'basicui']

root_dir = "C:\Program Files (x86)\Steam"
for dir_path, dir_names, file_names in os.walk(root_dir):
    for file_name in file_names:
        file_path = os.path.join(dir_path, file_name)
        file_ext = os.path.splitext(file_path)[1]
        if any(each_reduction in file_name for each_reduction in fileReduction):
            continue
        if file_ext in extensions:
            sheet = sheets[file_ext]
            sheet.append([file_path])

# Save the workbook
workbook.save("index.xlsx")
